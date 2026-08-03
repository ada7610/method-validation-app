import io
import openpyxl
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pandas as pd
import numpy as np
import streamlit as st

# ==========================================
# 🎨 Page & Sidebar Configuration
# ==========================================
st.set_page_config(
    page_title="Method Validation System | Abdulrahman Alamri",
    page_icon="🧪",
    layout="wide",
)

with st.sidebar:
    st.header("👑 Developer Info")
    st.markdown("""
    **Developed & Designed By:**  
    👨‍🔬 **Abdulrahman Alamri**  
    
    **Specialization:**  
    🔬 Quality Control & Method Validation Specialist  
    
    ---
    *All Rights Reserved © 2026*
    """)


# ==========================================
# 📊 Excel Generation Function
# ==========================================
def generate_validation_excel(
    calib_df, level1_df, test_title, unit_str, target_conc, t_val, std_purity
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validation Report"
    ws.views.sheetView[0].showGridLines = True

    COLOR_GREEN_HEADER = "C6EFCE"
    COLOR_BLUE_HEADER = "8EA9DB"
    COLOR_ORANGE_HEADER = "F4B084"
    COLOR_GRAY_NOTE = "D9D9D9"

    COLOR_PURPLE_HEADER = "7030A0"
    COLOR_PURPLE_SUB = "D9E1F2"

    font_main_title = Font(name="Calibri", size=14, bold=True)
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_regular = Font(name="Calibri", size=10)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    align_left = Alignment(horizontal="left", vertical="center")

    thin_border_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(
        left=thin_border_side, right=thin_border_side,
        top=thin_border_side, bottom=thin_border_side
    )

    # 1. Main File Title
    title_text = f"Calculation of Validation for {test_title}" if test_title else "Calculation of Validation"
    ws.merge_cells("A1:K1")
    ws["A1"] = title_text
    ws["A1"].font = font_main_title
    ws["A1"].fill = PatternFill("solid", fgColor=COLOR_GREEN_HEADER)
    ws["A1"].alignment = align_center

    # 2. Calibration Table
    ws.merge_cells("A4:C4")
    ws["A4"] = "Calibration STD"
    ws["A4"].font = font_bold
    ws["A4"].fill = PatternFill("solid", fgColor=COLOR_BLUE_HEADER)
    ws["A4"].alignment = align_center

    unit_header = f"concentration ({unit_str})" if unit_str else "concentration"
    ws["A5"] = "Level"
    ws["B5"] = unit_header
    ws["C5"] = "area"

    for col in ["A", "B", "C"]:
        ws[f"{col}5"].font = font_bold
        ws[f"{col}5"].fill = PatternFill("solid", fgColor=COLOR_ORANGE_HEADER)
        ws[f"{col}5"].alignment = align_center

    start_cal_row = 6
    for idx, row in calib_df.iterrows():
        r = idx + start_cal_row
        ws[f"A{r}"] = str(row.get("Level", ""))
        ws[f"B{r}"] = float(row.get("Concentration", 0)) if pd.notnull(row.get("Concentration")) else 0.0
        ws[f"C{r}"] = float(row.get("Area", 0)) if pd.notnull(row.get("Area")) else 0.0
        ws[f"B{r}"].number_format = "0.0000"
        ws[f"C{r}"].number_format = "0.0000"

    end_cal_row = max(len(calib_df) + start_cal_row - 1, start_cal_row)

    # 3. Grubbs Critical Values Table & Reference
    ws.merge_cells("J4:K4")
    ws["J4"] = "Critical values of G (P=0.05)"
    ws["J4"].font = font_white_bold
    ws["J4"].fill = PatternFill("solid", fgColor=COLOR_PURPLE_HEADER)
    ws["J4"].alignment = align_center

    ws["J5"] = "Sample size"
    ws["K5"] = "Critical value"
    for col_ref in ["J5", "K5"]:
        ws[col_ref].font = font_bold
        ws[col_ref].fill = PatternFill("solid", fgColor=COLOR_PURPLE_SUB)
        ws[col_ref].alignment = align_center
        ws[col_ref].border = thin_border

    grubbs_table = [
        (3, 1.155), (4, 1.481), (5, 1.715), (6, 1.887),
        (7, 2.020), (8, 2.126), (9, 2.215), (10, 2.290)
    ]

    last_grubbs_row = 5
    for row_idx, (n_val, g_crit) in enumerate(grubbs_table, 6):
        ws[f"J{row_idx}"] = n_val
        ws[f"K{row_idx}"] = g_crit
        ws[f"J{row_idx}"].alignment = align_center
        ws[f"K{row_idx}"].number_format = "0.0000"
        for c in ["J", "K"]:
            ws[f"{c}{row_idx}"].font = font_regular
            ws[f"{c}{row_idx}"].border = thin_border
        last_grubbs_row = row_idx

    # Reference text below the table
    ref_row = last_grubbs_row + 1
    ws.merge_cells(f"J{ref_row}:K{ref_row+1}")
    ws[f"J{ref_row}"] = "Ref: Miller& Miller, Statistics and chemometrics for analytical chemistry, 6th edition."
    ws[f"J{ref_row}"].font = Font(name="Calibri", size=9, italic=True, bold=True)
    ws[f"J{ref_row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 4. Scatter Chart
    chart = ScatterChart()
    chart.title = None

    chart.x_axis.title = None
    chart.y_axis.title = None

    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.crosses = "autoZero"
    chart.y_axis.crosses = "autoZero"
    chart.x_axis.tickLblPos = "nextTo"
    chart.y_axis.tickLblPos = "nextTo"

    chart.x_axis.number_format = "0.00"
    chart.y_axis.number_format = "0.00"

    chart.graphicalProperties = GraphicalProperties()
    chart.graphicalProperties.noFill = True

    chart.plot_area.graphicalProperties = GraphicalProperties()
    chart.plot_area.graphicalProperties.noFill = True

    chart.x_axis.majorGridlines = ChartLines()
    chart.x_axis.majorGridlines.graphicalProperties = GraphicalProperties()
    chart.x_axis.majorGridlines.graphicalProperties.line = LineProperties(solidFill="D9D9D9")

    chart.y_axis.majorGridlines = ChartLines()
    chart.y_axis.majorGridlines.graphicalProperties = GraphicalProperties()
    chart.y_axis.majorGridlines.graphicalProperties.line = LineProperties(solidFill="D9D9D9")

    xvalues = Reference(ws, min_col=2, min_row=start_cal_row, max_row=end_cal_row)
    yvalues = Reference(ws, min_col=3, min_row=start_cal_row, max_row=end_cal_row)

    series = Series(yvalues, xvalues, title_from_data=False)
    series.marker.symbol = "circle"
    series.marker.size = 6
    series.graphicalProperties.line.noFill = True

    series.dataLabels = DataLabelList()
    series.dataLabels.showVal = False
    series.dataLabels.showCatName = False
    series.dataLabels.showSerName = False

    trendline = openpyxl.chart.trendline.Trendline(
        trendlineType="linear", dispEq=True, dispRSqr=True
    )
    try:
        trendline.graphicalProperties = GraphicalProperties()
        trendline.graphicalProperties.line = LineProperties(prstDash="sysDot", cmpd="sng")
    except Exception:
        pass

    series.trendline = trendline
    chart.series.append(series)
    chart.legend = None
    
    chart.width = 13
    chart.height = 7
    ws.add_chart(chart, "F3")

    # 5. RSQ, t-value, Spiked Level
    rsq_row = end_cal_row + 2
    tval_row = rsq_row + 1
    spiked_row = tval_row + 1

    ws[f"A{rsq_row}"] = "RSQ"
    ws[f"B{rsq_row}"] = f"=RSQ(C{start_cal_row}:C{end_cal_row}, B{start_cal_row}:B{end_cal_row})"

    ws[f"A{tval_row}"] = "t-value (t test)"
    ws[f"B{tval_row}"] = float(t_val)

    ws[f"A{spiked_row}"] = "Spiked Level"
    ws[f"B{spiked_row}"] = float(target_conc)

    for r in [rsq_row, tval_row, spiked_row]:
        ws[f"A{r}"].font = font_bold
        ws[f"A{r}"].fill = PatternFill("solid", fgColor=COLOR_ORANGE_HEADER)
        ws[f"A{r}"].alignment = align_left

        ws[f"B{r}"].font = font_bold
        ws[f"B{r}"].fill = PatternFill(fill_type=None)
        ws[f"B{r}"].alignment = align_center
        ws[f"B{r}"].number_format = "0.0000"

    # 6. Level 1 Table
    l1_title_row = spiked_row + 2
    l1_header_row = l1_title_row + 1
    start_sample_row = l1_header_row + 1

    ws.merge_cells(f"A{l1_title_row}:E{l1_title_row}")
    ws[f"A{l1_title_row}"] = f"Level 1 ({unit_str})" if unit_str else "Level 1"
    ws[f"A{l1_title_row}"].font = font_bold
    ws[f"A{l1_title_row}"].fill = PatternFill("solid", fgColor=COLOR_BLUE_HEADER)
    ws[f"A{l1_title_row}"].alignment = align_center

    headers_l1 = ["Samples name", unit_header, "Recovery %", "Outlier (Z-Score)", "Outlier Status"]
    cols_l1 = ["A", "B", "C", "D", "E"]
    for c, h in zip(cols_l1, headers_l1):
        ws[f"{c}{l1_header_row}"] = h
        ws[f"{c}{l1_header_row}"].font = font_bold
        ws[f"{c}{l1_header_row}"].fill = PatternFill("solid", fgColor=COLOR_ORANGE_HEADER)
        ws[f"{c}{l1_header_row}"].alignment = align_center

    for idx, row in level1_df.iterrows():
        r = idx + start_sample_row
        ws[f"A{r}"] = str(row.get("Sample Name", ""))
        sample_conc = float(row.get("Concentration", 0)) if pd.notnull(row.get("Concentration")) else 0.0
        ws[f"B{r}"] = sample_conc
        ws[f"B{r}"].number_format = "0.0000"

    end_sample_row = max(len(level1_df) + start_sample_row - 1, start_sample_row)

    # 7. Statistics & Calculations
    stats_start_row = end_sample_row + 2
    mean_row = stats_start_row
    rec_row = stats_start_row + 1
    sd_row = stats_start_row + 2
    rsd_row = stats_start_row + 3

    stats_labels = [
        ("Mean", f"=AVERAGE(B{start_sample_row}:B{end_sample_row})"),
        ("Recovery %", f"=AVERAGE(C{start_sample_row}:C{end_sample_row})"),
        ("Standerd Deviation", f"=STDEV(B{start_sample_row}:B{end_sample_row})"),
        ("RSD %", f"=IF(B{mean_row}=0, 0, (B{sd_row}/B{mean_row})*100)"),
        ("LOD", f"=B{tval_row}*B{sd_row}"),
        ("LOQ", f"=10*B{sd_row}"),
    ]

    for i, (label, formula) in enumerate(stats_labels, start=stats_start_row):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = font_bold
        ws[f"A{i}"].fill = PatternFill("solid", fgColor=COLOR_BLUE_HEADER)
        ws[f"B{i}"] = formula
        ws[f"B{i}"].number_format = "0.0000"

    # Formulas for Recovery and Outlier (Z-Score) & Status
    for r in range(start_sample_row, end_sample_row + 1):
        ws[f"C{r}"] = f"=IF(B{spiked_row}=0, 0, (B{r}/B{spiked_row})*100)"
        ws[f"D{r}"] = f"=ABS(B{r}-B${mean_row})/B${sd_row}"
        ws[f"E{r}"] = f'=IF(D{r}>VLOOKUP(COUNT(B${start_sample_row}:B${end_sample_row}), J$6:K$13, 2, FALSE), "Outlier", "Normal")'
        
        ws[f"C{r}"].number_format = "0.0000"
        ws[f"D{r}"].number_format = "0.0000"
        ws[f"E{r}"].alignment = align_center

    ws.merge_cells(f"F{start_sample_row}:F{end_sample_row}")
    ws[f"F{start_sample_row}"] = "Any value higher than the critical value in the table is consider outlier"
    ws[f"F{start_sample_row}"].font = Font(name="Calibri", size=9, bold=True)
    ws[f"F{start_sample_row}"].fill = PatternFill("solid", fgColor=COLOR_GRAY_NOTE)
    ws[f"F{start_sample_row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 8. Measurement Uncertainty Table
    unc_header_row = l1_header_row
    unc_start_row = start_sample_row

    ws.merge_cells(f"H{unc_header_row}:I{unc_header_row}")
    ws[f"H{unc_header_row}"] = "Measurment uncertainty"
    ws[f"H{unc_header_row}"].font = font_bold
    ws[f"H{unc_header_row}"].fill = PatternFill("solid", fgColor=COLOR_ORANGE_HEADER)
    ws[f"H{unc_header_row}"].alignment = align_center

    ws.merge_cells(f"H{unc_start_row}:I{unc_start_row}")
    ws[f"H{unc_start_row}"] = "Level 1"
    ws[f"H{unc_start_row}"].font = font_bold
    ws[f"H{unc_start_row}"].fill = PatternFill("solid", fgColor=COLOR_BLUE_HEADER)
    ws[f"H{unc_start_row}"].alignment = align_center

    purity_row = unc_header_row - 1
    ws[f"H{purity_row}"] = "Standard Purity"
    ws[f"H{purity_row}"].font = font_bold
    purity_val = std_purity / 100.0 if std_purity > 1.0 else std_purity
    ws[f"I{purity_row}"] = purity_val
    ws[f"I{purity_row}"].number_format = "0.0000"

    uA_row = unc_start_row + 1
    uB_row = uA_row + 1
    uC_row = uB_row + 1
    uD_row = uC_row + 1
    uComb_row = uD_row + 1
    uExp_row = uComb_row + 1

    unc_labels = [
        (uA_row, "uA", f"=B{rsd_row}/100"),
        (uB_row, "uB", f"=ABS(0.5*(1 - (B{rec_row}/100)))/SQRT(3)"),
        (uC_row, "uC", f"=0.5*(1 - I{purity_row})/SQRT(3)"),
        (uD_row, "uD", f"=1 - SQRT(B{rsq_row})"),
        (uComb_row, "u combiend", f"=SQRT(I{uA_row}^2 + I{uB_row}^2 + I{uC_row}^2 + I{uD_row}^2)"),
        (uExp_row, "U expanded", f"=2*I{uComb_row}"),
    ]

    for r_num, u_name, u_formula in unc_labels:
        ws[f"H{r_num}"] = u_name
        ws[f"I{r_num}"] = u_formula
        ws[f"I{r_num}"].number_format = "0.0000"
        if u_name in ["u combiend", "U expanded"]:
            ws[f"H{r_num}"].font = font_bold
            ws[f"I{r_num}"].font = font_bold

    # Column Widths Setup
    column_widths = {
        "A": 22, "B": 24, "C": 18, "D": 18, "E": 18,
        "F": 25, "H": 22, "I": 18, "J": 25, "K": 25
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ==========================================
# ⚙️ Streamlit User Interface (English Only)
# ==========================================
st.title("🧪 Analytical Method Validation System")

col_header1, col_header2 = st.columns(2)
with col_header1:
    test_name = st.text_input("Test / Analysis Name (e.g. Benzo a pyrene)", "Benzo a pyrene")
with col_header2:
    conc_unit = st.text_input("Concentration Unit", "ppm")

st.divider()

st.subheader("📌 Calibration Standard Table (Calibration STD)")
num_calib_levels = st.number_input(
    "Number of Calibration Levels", min_value=1, max_value=30, value=6, step=1, key="calib_num_input"
)

default_concs = [0.5, 1.0, 5.0, 10.0, 50.0, 100.0]
default_areas = [10.5, 25.0, 100.2, 180.4, 720.5, 1550.0]

calib_data = [
    {
        "Level": f"STD {i+1}",
        "Concentration": default_concs[i] if i < len(default_concs) else 0.0000,
        "Area": default_areas[i] if i < len(default_areas) else 0.0000
    }
    for i in range(int(num_calib_levels))
]

valid_std_raw = st.data_editor(
    pd.DataFrame(calib_data), num_rows="dynamic", key=f"calib_table_editor_{num_calib_levels}", use_container_width=True
)

valid_std = valid_std_raw.dropna(how="all").copy()
if "Level" in valid_std.columns:
    valid_std["Level"] = valid_std["Level"].fillna("")
if "Concentration" in valid_std.columns:
    valid_std["Concentration"] = pd.to_numeric(valid_std["Concentration"], errors="coerce").fillna(0.0)
if "Area" in valid_std.columns:
    valid_std["Area"] = pd.to_numeric(valid_std["Area"], errors="coerce").fillna(0.0)

st.divider()

st.subheader("📋 Samples & Inputs Table (Level 1)")
col_input1, col_input2, col_input3 = st.columns(3)
with col_input1:
    target_conc = st.number_input("Spiked Level", value=10.0000, min_value=0.0, format="%.4f")
with col_input2:
    t_val = st.number_input("t-statistic value", value=2.5710, format="%.4f")
with col_input3:
    std_purity = st.number_input("Standard Purity", value=0.9900, min_value=0.0, max_value=100.0, format="%.4f")

num_samples = st.number_input(
    "Number of Replicates / Samples", min_value=1, max_value=30, value=6, step=1, key="samples_num_input"
)

sample_data = [
    {"Sample Name": f"Sample {i+1}", "Concentration": 0.0000} for i in range(int(num_samples))
]

edited_samples_raw = st.data_editor(
    pd.DataFrame(sample_data), num_rows="dynamic", key=f"samples_table_editor_{num_samples}", use_container_width=True
)

edited_samples = edited_samples_raw.dropna(how="all").copy()
if "Sample Name" in edited_samples.columns:
    edited_samples["Sample Name"] = edited_samples["Sample Name"].fillna("")
if "Concentration" in edited_samples.columns:
    edited_samples["Concentration"] = pd.to_numeric(edited_samples["Concentration"], errors="coerce").fillna(0.0)

st.divider()

try:
    calib_export = (
        valid_std[["Level", "Concentration", "Area"]]
        if not valid_std.empty
        else pd.DataFrame(columns=["Level", "Concentration", "Area"])
    )

    excel_file = generate_validation_excel(
        calib_df=calib_export,
        level1_df=edited_samples,
        test_title=test_name,
        unit_str=conc_unit,
        target_conc=target_conc,
        t_val=t_val,
        std_purity=std_purity,
    )

    st.download_button(
        label="📥 Download Final Validation Excel Report",
        data=excel_file,
        file_name="Method_Validation_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
except Exception as e:
    st.error(f"An error occurred while generating the Excel file: {e}")

st.caption("---")
st.caption("Developed with ❤️ by **Abdulrahman Alamri** | All Rights Reserved © 2026")
